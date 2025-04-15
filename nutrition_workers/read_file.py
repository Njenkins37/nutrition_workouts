import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import declarative_base
from .nutrition_tables import Food, Workouts
from .database_interface import DatabaseInit

Base = declarative_base()


class DatabaseInterface(DatabaseInit):
    def __init__(self):
        """
        The DatabaseInterface class converts the Excel file information into a dictionary to be inserted into the database.
        The Excel file is located in the Files location.
        """
        super().__init__()
        self.file_name = './Files/nutrition_workouts.xlsx'
        self.workouts_sheet = 'workouts'
        self.foods_sheet = 'foods'
        self.diet_sheet = 'diet'
        self.log_sheet = 'log'
        self.mood_sheet = 'mood'
        self.food_map = self.get_food_id()
        self.workout_map = self.get_workout_id()
        self.foods_dict = self.get_data(self.foods_sheet)
        self.workout_dict = self.get_data(self.workouts_sheet)
        self.diet_dict = self.get_diet_log('diet')
        self.log_dict = self.get_diet_log('log')
        self.mood_dict = self.get_data(self.mood_sheet)

    def get_data(self, sheet: str) -> list:
        """
        Calls the get_columns fucntion to get the mapping of data for the both sheets in the nutrition excel file
        :return: The data in a dictionary format to be inserted in the database
        """
        data = self.get_dict(pd.read_excel(self.file_name, sheet_name=sheet))

        is_clear = self.clear_sheet_but_keep_header(sheet)
        if is_clear:
            print(f'Data retreived from {sheet} and removed from excel file')
        else:
            print(f'There was no data to retrieve from {sheet}')

        return data

    @staticmethod
    def get_dict(df: pd.DataFrame) -> list:
        """

        :param df: The dataframe resulted from reading the passed excel file and the sheet name
        :return: the dataframe in a dictionary format to be inserted into the dictionary
        """
        if 'muscles' in df.columns:
            df['muscles'] = df['muscles'].apply(
                lambda x: [m.strip().title() for m in x.split(',')] if isinstance(x, str) else x
            )
        if 'name' in df.columns:
            df['name'] = df['name'].apply(lambda x: x.title())

        return [
            {column: row[column] for column in df.columns}
            for _, row in df.iterrows()
        ]

    def clear_sheet_but_keep_header(self, sheet_name: str) -> bool:
        """

        :param sheet_name: The sheet name to delete from
        :return: Boolean. Used in check to determine which log message to print
        """

        # Removes content from the workbook so that it's not reuploaded into the database.
        workbook = load_workbook(self.file_name)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)
                workbook.save(self.file_name)
                return True
            else:
                return False
        return False

    def get_food_id(self):
        """
        Gets the food mapping dictionary based on food name
        :return: the food mapping dictionary
        """
        session = self.Session()
        food_map = {food.name: food.id for food in session.query(Food).all()}
        session.close()
        return food_map

    def get_workout_id(self):
        """
        Gets the workout mapping dictionary based on workout name
        :return: the workout dictionary
        """
        session = self.Session()
        workout_map = {workout.name: workout.id for workout in session.query(Workouts).all()}
        session.close()
        return workout_map

    def get_diet_log(self, param):
        """
        Reads the diet and logs sheets of the excel file and adds it to the database. Also calls the clear Excel file
        function after retrieving the data.
        :param param: Used to separate which mapping to return
        :return: the data dictionary used to insert the data into database
        """

        # Find the mapping
        if param == 'diet':
            df = pd.read_excel(self.file_name, self.diet_sheet)
            df['name'] = df['name'].apply(lambda x: x.strip().title())  # format name in title to match database
            df['food_id'] = df['name'].map(self.food_map)
            for idx, row in df[df['food_id'].isnull()].iterrows():
                best_match = self.lcs('food', row['name'])  # Get best matching food name
                if best_match:
                    df.at[idx, 'food_id'] = self.food_map.get(best_match)
        elif param == 'log':
            df = pd.read_excel(self.file_name, self.log_sheet)
            df['name'] = df['name'].apply(lambda x: x.strip().title())  # format name in title to match database
            df['workout_id'] = df['name'].map(self.workout_map)
            for idx, row in df[df['workout_id'].isnull()].iterrows():
                best_match = self.lcs('workout', row['name'])  # Get best matching food name
                if best_match:
                    df.at[idx, 'workout_id'] = self.workout_map.get(best_match)

        df.drop('name', axis=1, inplace=True)
        data = self.get_dict(df)
        # If is_clear it will print that the data was retreived
        is_clear = self.clear_sheet_but_keep_header(param)
        if is_clear:
            print(f'Data retrieved from {param} sheet and sheet cleared')
        else:
            print(f'There was no data to retrieve from {param}')
        return data

    def lcs(self, param, word):

        if param == 'food':
            iterable = self.food_map.keys()
        elif param == 'workout':
            iterable = self.workout_map.keys()
        else:
            iterable=None

        max_item = None
        max_score = -1

        for item in iterable:

            m, n = len(item), len(word)
            T = [[0] * (n + 1) for _ in range(m + 1)]
            for i in range(1, m + 1):
                for j in range(1, n + 1):
                    if item[i - 1] == word[j - 1]:
                        T[i][j] = T[i - 1][j - 1] + 1
                    else:
                        T[i][j] = max(T[i - 1][j], T[i][j - 1])
            lcs_score = T[m][n]
            if lcs_score > max_score:
                max_score = lcs_score
                max_item = item
            return max_item



if __name__ == '__main__':
    data = DatabaseInterface()
    print(data.diet_dict)
